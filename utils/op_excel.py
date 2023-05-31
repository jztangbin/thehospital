import datetime
from copy import copy
import datetime
from time import strftime
import openpyxl
import yaml
import time



f = open('/users/seasea/thehospital/theConfig.yaml')#指定yaml文件路径
data = yaml.load(f.read(),Loader=yaml.FullLoader)  #读配置
wb = openpyxl.load_workbook(data['excel']['local'])  #实例化
newwb = copy(wb)
sheet = newwb.active
savetime = datetime.datetime.now()
reportwb = openpyxl.load_workbook(data['excel']['reportlocal'])
rewb = copy(reportwb)
resheet = rewb.active




class exceloperate(object):  #获取身份证
    def getcarnumber(self,row):
        carnumber = (sheet.cell(row, 4).value)
        return carnumber

    def reportcarnum(self,row):
         canumber = (resheet.cell(row,6).value)
         return canumber

    def updatedata(self,row,key):  #新建表格存储数据
        sheet['E'+str(row)] = key
        #newwb.save(data['excel']['sqvelocal']+"Rankingfilter"+savetime.strftime("%Y-%m-%d")+".xlsx")
        return print(sheet.cell(row,3).value,"的数据更新为",sheet.cell(row,5).value)


if  __name__ =='__main__':
    exceloperate().reportcarnum(2)
    #exceloperate().getcarnumber(6)
